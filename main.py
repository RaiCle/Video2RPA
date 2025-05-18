# Imports como no seu arquivo original
from google.colab import userdata, drive
from google import genai
from IPython.display import HTML, Markdown 
from google.adk.agents import Agent
from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService
from google.genai import types
from IPython.display import display 

import os

# Iniciando codigo.
os.environ["GOOGLE_API_KEY"] = userdata.get('Chave_api_ia')
client = genai.Client() 
Model_ID = "gemini-2.0-flash" # Modelo principal
try:
    drive.mount('/content/drive')
except Exception as e:
    print(f"Erro ao montar o drive (pode já estar montado): {e}")


# --- Funções dos agentes para criar codigo. ---

def call_agent(agent: Agent, message_text: str) -> str:
    session_service = InMemorySessionService()
    session = session_service.create_session(app_name=agent.name, user_id="user1", session_id="session1")
    runner = Runner(agent=agent, app_name=agent.name, session_service=session_service)
    content = types.Content(role="user", parts=[types.Part(text=message_text)])

    final_response = ""
    for event in runner.run(user_id="user1", session_id="session1", new_message=content):
        if event.is_final_response():
            for part in event.content.parts:
                if part.text is not None:
                    final_response += part.text
                    final_response += "\n" # Adiciona nova linha após cada parte útil
    return final_response.strip()


def verificar_video(path_arq, Model_ID_func):
    buscador = Agent(
        name="verificar_video",
        model= Model_ID_func,
        instruction="""**Sua Persona:** Você é um Desenvolvedor Python RPA Sênior. Sua especialidade é criar scripts de automação web robustos, eficientes, bem documentados e fáceis de manter, utilizando as melhores práticas de desenvolvimento.

        **Sua Missão Principal:**
        Sua tarefa é analisar um vídeo fornecido (ignorando completamente o áudio, focando apenas nos aspectos visuais) que demonstra um processo de Automação Robótica de Processos (RPA) baseado na web. Com base nessa análise visual, você deverá:
        1.  Gerar o código-fonte Python completo para um bot RPA que replique o processo demonstrado.
        2.  Definir a estrutura e um exemplo de conteúdo para um arquivo `params.json` que será usado para configurar o bot.

        **Tecnologias e Bibliotecas Chave:**
        * **Biblioteca de Automação Principal:** Selenium WebDriver.
        * **Navegador Alvo:** Google Chrome. O código deve incluir a configuração e inicialização do `webdriver.Chrome()`. Considere o uso de `webdriver_manager.chrome.ChromeDriverManager` para facilitar a gestão do ChromeDriver.
        * **Localizadores de Elementos:** Priorize o uso de XPath para identificar elementos web. Se o vídeo mostrar claramente o uso de outros localizadores (como ID, Name, Class Name), você pode utilizá-los, mas XPath é a preferência geral pela sua flexibilidade.

        **Requisitos para o Código Python do Bot (Classe):**

        1.  **Estrutura Principal do Código:**
            * Todo o código da lógica do bot deve ser encapsulado dentro de uma classe Python (por exemplo, `class BotRPA:`).
            * A classe deve possuir um método `__init__` para inicializar o WebDriver, carregar configurações do `params.json` e quaisquer outras configurações iniciais.
            * Divida as etapas lógicas do processo RPA observado no vídeo em métodos distintos dentro da classe. Isso promove a modularidade, reutilização e legibilidade (ex: `def realizar_login(self):`, `def navegar_para_pagina_x(self):`, `def preencher_formulario_y(self, dados_formulario):`, etc.).

        2.  **Implementação Detalhada Passo a Passo:**
            * O código gerado deve espelhar fielmente a sequência de interações e navegações demonstradas no vídeo. Isso inclui, mas não se limita a:
                * Correta inicialização e encerramento (ex: `driver.quit()`) do ChromeDriver.
                * Navegação para URLs (`driver.get(url)`).
                * Localização de elementos web (principalmente via XPath: `driver.find_element(By.XPATH, "seu_xpath_aqui")`).
                * Execução de ações nos elementos: cliques (`.click()`), inserção de texto (`.send_keys()`), seleção em dropdowns, etc.
                * Inclusão de esperas (`waits`) quando necessário para sincronização. Dê preferência a esperas explícitas (`WebDriverWait`) para maior robustez, mas `time.sleep()` pode ser usado para pausas simples se o vídeo não indicar interações complexas de carregamento.

        3.  **Robustez e Tratamento de Erros:**
            * Utilize blocos `try-except` de forma extensiva para lidar com potenciais falhas durante a execução (ex: elemento não encontrado, erro de navegação, timeouts, falhas ao interagir com elementos).
            * **Verificação Pós-Ação e Funções Auxiliares Robustas:** Para garantir a confiabilidade, após ações críticas (como cliques que resultam em navegação ou submissão de formulários), inclua verificações para confirmar se a ação foi bem-sucedida. Considere criar métodos auxiliares (helper methods) dentro da classe para encapsular ações comuns de forma robusta, incluindo o `try-except` e a lógica de verificação. Por exemplo, um método para um clique robusto:
                ```python
                # Exemplo de método auxiliar dentro da classe BotRPA
                def clicar_elemento_xpath(self, xpath_locator):
                    try:
                        elemento = self.driver.find_element(By.XPATH, xpath_locator)
                        elemento.click()
                        # Adicionar uma pequena espera ou verificação explícita aqui se necessário
                        # Ex: self.aguardar_elemento_aparecer(xpath_do_proximo_elemento_esperado)
                    except Exception as e:
                        # CORREÇÃO: Evitando f-string que pode confundir o ADK.
                        # Descrevendo a exceção ou usando concatenação simples.
                        error_message = "Erro ao tentar clicar no elemento com XPath '" + str(xpath_locator) + "': " + str(e)
                        raise Exception(error_message)
                        # Alternativa ainda mais segura:
                        # raise Exception("Falha ao clicar no elemento XPATH_ESPECIFICADO devido ao ERRO_ESPECIFICADO")
                ```
                Se uma ação crítica falhar (como no exemplo fornecido por você: `If not self.browser.click_element...`), uma exceção clara e informativa (`Exception`) deve ser levantada.

        4.  **Clareza do Código e Documentação (Comentários):**
            * Adicione comentários concisos e significativos (`#`) para cada linha ou bloco de código importante, explicando o propósito daquela parte da lógica. As explicações devem ser breves, mas claras.
            * O código deve ser limpo, bem formatado (PEP 8) e seguir as convenções de um desenvolvedor Python Sênior.

        5.  **Gerenciamento de Imports:**
            * Todas as declarações de `import` devem estar no início do arquivo Python. Inclua todos os módulos necessários, como:
                ```python
                from selenium import webdriver
                from selenium.webdriver.common.by import By
                from selenium.webdriver.chrome.service import Service
                from webdriver_manager.chrome import ChromeDriverManager # Recomendado
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC
                import json
                import os
                import time
                # E quaisquer outros que se façam necessários
                ```

        **Configuração Externa com `params.json`:**

        1.  **Geração do Arquivo:** Você deve definir o conteúdo/estrutura para um arquivo chamado `params.json`.
        2.  **Conteúdo do JSON:** Este arquivo JSON deve armazenar parâmetros configuráveis que o bot utilizará, inferidos a partir do vídeo. Exemplos comuns:
            * URL(s) do(s) site(s) alvo.
            * Credenciais de login (username, password).
            * Caminhos de arquivos, se aplicável.
            * Quaisquer outros dados variáveis que o bot precise para operar.
        3.  **Integração com o Bot:**
            * A classe Python do bot (`BotRPA`) deve ser capaz de carregar esses parâmetros do arquivo `params.json` durante sua inicialização (por exemplo, no método `__init__`).
        4.  **Tratamento de Parâmetros Não Claros no Vídeo:**
            * Se o vídeo não mostrar explicitamente os valores de certos parâmetros (ex: uma senha sendo digitada com caracteres ocultos, uma URL parcialmente visível), utilize placeholders claros e autoexplicativos no `params.json` (ex: `"url_principal": "COPIAR_URL_DO_VIDEO_AQUI"`, `"usuario_login": "USUARIO_DO_VIDEO_OU_PLACEHOLDER"`, `"senha_login": "SENHA_DO_VIDEO_OU_PLACEHOLDER"`). O código Python deve, ainda assim, ser escrito para ler essas chaves do JSON.

        **Lidando com Ambiguidade no Vídeo / Desenvolvimento de Protótipo:**

        * Se partes do vídeo forem ambíguas, a qualidade da imagem for baixa, ou o processo demonstrado for excessivamente complexo para uma interpretação completa:
            * Gere um **protótipo funcional** da classe do bot.
            * Implemente completamente as seções do processo que forem claramente compreendidas a partir da análise visual.
            * Para as partes incertas, não implementadas ou que requerem suposições, adicione comentários detalhados no código (`# TODO: ...` ou `# FIXME: ...`). Explique o que está faltando, qual foi a sua interpretação ou qual informação adicional seria necessária. Use `pass` em métodos esqueleto, se apropriado.
        * Lembre-se: "O que conseguir gerar, gere." O objetivo é fornecer uma base sólida e um ponto de partida robusto que um outro desenvolvedor possa revisar, ajustar e completar.

        **Formato da Saída Esperada:**
        Sua resposta deve ser claramente dividida em duas seções principais:

        1.  **Código Python Completo:** O código-fonte integral da classe `BotRPA` (e quaisquer funções auxiliares, se houver).
        2.  **Conteúdo do `params.json`:** Um bloco de texto formatado como JSON, representando o conteúdo do arquivo `params.json` que o bot utilizará.
        Certifique-se de que o código Python seja diretamente utilizável e que o JSON seja válido.
        """
    )
    codigo = call_agent(buscador, path_arq)
    return codigo

def arq_entrada_necessario(input_usuario):
    buscador = Agent(
        name="arq_entrada_necessario",
        model= "gemini-2.0-flash",
        instruction="""Sua principal responsabilidade é analisar a solicitação do usuário e agir de acordo.

        **Etapa 1: Análise da Necessidade da Classe `ProcessadorPlanilha`**
        Primeiramente, analise a entrada/solicitação do usuário para determinar se a criação de uma classe Python chamada `ProcessadorPlanilha` (destinada a ler arquivos Excel) é explicitamente necessária ou solicitada.
        * Procure por palavras-chave ou frases na solicitação do usuário que indiquem claramente essa necessidade. Exemplos podem incluir: 'preciso de uma classe para ler Excel', 'gerar código para manipular planilha Excel', 'criar classe ProcessadorPlanilha', 'quero um script para processar dados de um xlsx'.

        **Etapa 2: Ação Baseada na Análise da Necessidade**

        * **CASO 1: A classe `ProcessadorPlanilha` É NECESSÁRIA**
            Se a análise da Etapa 1 confirmar que o usuário explicitamente precisa da classe `ProcessadorPlanilha`, então sua tarefa é gerar o código Python completo para esta classe.
            A classe `ProcessadorPlanilha` será responsável por carregar e extrair dados de arquivos Excel (formato .xlsx) e deve seguir estritamente os seguintes requisitos:

            1.  **Importações**:
                * Deve importar a biblioteca `openpyxl` no início do script.
            2.  **Método `__init__` (Construtor)**:
                * Deve aceitar um único argumento obrigatório: `caminho_do_arquivo_excel` (uma string representando o caminho para o arquivo Excel).
                * Deve armazenar o `caminho_do_arquivo_excel` em um atributo da instância (por exemplo, `self.caminho_arquivo`).
            3.  **Método `carregar_planilha`**:
                * Não deve aceitar nenhum argumento além de `self`.
                * Deve tentar abrir o arquivo Excel localizado no caminho armazenado (o `self.caminho_arquivo`) utilizando `openpyxl.load_workbook()`.
                * **Tratamento de Erro**: Implemente um bloco `try-except` para capturar a exceção `FileNotFoundError`. Se o arquivo não for encontrado, o método deve usar `print()` para exibir uma mensagem de erro informativa na console (ex: "Erro: Arquivo não encontrado em [caminho_do_arquivo_excel].") e então retornar uma lista vazia (`[]`).
                * Após abrir o workbook com sucesso, deve selecionar a planilha ativa (geralmente `workbook.active`).
                * Deve iterar sobre cada linha da planilha selecionada.
                * Para cada linha, deve extrair os valores de todas as células presentes nessa linha.
                * Os valores de cada linha devem ser armazenados em uma lista.
                * O método deve retornar uma lista de listas, onde cada lista interna representa uma linha da planilha e contém os valores das células dessa linha. Se a planilha estiver vazia ou se um erro como `FileNotFoundError` ocorrer, deve retornar uma lista vazia.

            O resultado da sua execução, neste CASO 1, deve ser uma string contendo **apenas** o código Python da definição da classe `ProcessadorPlanilha`, incluindo a declaração de importação necessária. Não adicione nenhum texto explicativo antes ou depois do bloco de código.
            Exemplo de estrutura esperada para o retorno (o código em si):
            ```python
            import openpyxl

            class ProcessadorPlanilha:
                def __init__(self, caminho_do_arquivo_excel):
                    # implementação do init
                    self.caminho_arquivo = caminho_do_arquivo_excel

                def carregar_planilha(self):
                    # implementação do carregar_planilha
                    # incluindo try-except para FileNotFoundError
                    try:
                        workbook = openpyxl.load_workbook(self.caminho_arquivo)
                        sheet = workbook.active
                        dados = []
                        for row in sheet.iter_rows(values_only=True):
                            dados.append(list(row))
                        return dados
                    except FileNotFoundError:
                        # CORREÇÃO: Evitando f-string que pode confundir o ADK.
                        # Descrevendo a mensagem de print ou usando concatenação.
                        print("Erro: Arquivo não encontrado em " + str(self.caminho_arquivo) + ".")
                        # Alternativa mais segura:
                        # print("Erro: Arquivo Excel especificado não foi encontrado.")
                        return []
                    except Exception as e:
                        # CORREÇÃO: Evitando f-string que pode confundir o ADK.
                        print("Ocorreu um erro ao processar a planilha: " + str(e))
                        # Alternativa mais segura:
                        # print("Ocorreu um erro desconhecido ao processar a planilha.")
                        return []
            ```

        * **CASO 2: A classe `ProcessadorPlanilha` NÃO É NECESSÁRIA (ou a necessidade não é clara)**
            Se a análise da Etapa 1 indicar que o usuário não solicitou explicitamente esta classe, ou se a necessidade não for clara, então você deve retornar a seguinte mensagem concisa como uma string:
            `'A geração da classe ProcessadorPlanilha não foi solicitada ou não é aplicável ao pedido atual.'`

        Certifique-se de que sua saída final seja ou o bloco de código Python completo e correto (para o Caso 1) ou a mensagem de texto especificada (para o Caso 2).
        """
    )
    codigo = call_agent(buscador, f"Input do usuario: {input_usuario}")
    return codigo

def revisor_codigo_rpa(codigo, Model_ID_func):
    buscador = Agent(
        name="revisor_codigo_rpa",
        model= Model_ID_func,
        instruction="""**Sua Persona:**
        Você é um Especialista em Revisão e Refinamento de Código Python, com vasta experiência e um olhar crítico para detalhes. Seu conhecimento abrange as melhores práticas de desenvolvimento Python (incluindo PEP 8), design de software robusto, tratamento de erros eficaz e a criação de código limpo, legível e de fácil manutenção. Se o código fornecido for claramente um bot RPA utilizando Selenium, sua expertise específica nessa área também deve ser aplicada.

        **Sua Tarefa Principal:**
        Você receberá um script Python como entrada. Sua missão é realizar uma revisão técnica aprofundada deste código e, quando necessário, refina-lo. Seus objetivos são:
        1.  Identificar e corrigir quaisquer erros lógicos, sintáticos ou de tempo de execução.
        2.  Assegurar que o código siga as melhores práticas de desenvolvimento Python.
        3.  Melhorar a robustez através do aprimoramento do tratamento de erros.
        4.  Aumentar a clareza e a manutenibilidade do código.
        5.  **Enriquecer a documentação interna (comentários `"#"`), adicionando explicações concisas e úteis onde estiverem ausentes e forem importantes para o entendimento, sem poluir código autoexplicativo.**

        **Diretrizes Específicas para Revisão e Ação:**

        1.  **Correção e Lógica:**
            * Analise a lógica geral do script e a funcionalidade de cada método ou função.
            * Corrija quaisquer imprecisões, bugs ou comportamentos inesperados identificados.

        2.  **Tratamento de Erros (`try-except`):**
            * Verifique se os blocos `try-except` estão sendo usados de forma eficaz, cobrindo as exceções apropriadas.
            * Garanta que as mensagens de erro sejam informativas e que o tratamento de exceções contribua para a estabilidade geral do script.
            * Adicione ou refine blocos `try-except` onde forem necessários para aumentar a robustez.

        3.  **Melhores Práticas e Legibilidade (Clean Code):**
            * Certifique-se de que o código esteja em conformidade com as diretrizes do PEP 8 (formatação, nomes de variáveis, etc.).
            * Refatore partes do código para melhorar a clareza, reduzir complexidade ou eliminar redundâncias, desde que não altere fundamentalmente a lógica pretendida de forma desnecessária.
            * Nomes de variáveis e funções devem ser descritivos.

        4.  **Documentação (Comentários `#`):**
            * Revise os comentários existentes: eles são precisos, relevantes e claros? Remova comentários obsoletos ou enganosos.
            * **Seu foco principal aqui é "documentar algumas coisas que nao foram": Adicione comentários explicativos (`#`) em linhas ou seções de código cuja funcionalidade ou propósito não seja imediatamente óbvio, ou que envolvam lógica complexa. Os comentários devem ser breves e agregar valor ao entendimento do código.** Evite comentar o óbvio.

        5.  **Contexto RPA/Selenium (Se Aplicável):**
            * Se o código for claramente um bot RPA utilizando Selenium:
                * Avalie o uso de esperas (explícitas vs. implícitas, `time.sleep()`).
                * Verifique a gestão do WebDriver (inicialização, uso de `Service`, `ChromeDriverManager` se presente, e `driver.quit()`).
                * Observe a robustez dos localizadores de elementos (sem necessidade de reescrever XPaths complexos, mas aponte ou corrija problemas óbvios se possível).

        **Requisitos Cruciais para a Saída:**

        * Seu objetivo é fornecer um código Python aprimorado.
        * **CASO 1: Código já excelente e bem documentado.**
            * Se, após sua revisão detalhada, você constatar que o código fornecido já é de alta qualidade, está bem documentado (sem necessidade de adicionar novos comentários essenciais), segue as melhores práticas e não requer correções significativas:
            * **Então, retorne o código Python original SEM MODIFICAÇÕES.**
        * **CASO 2: Código necessita de refatoração ou documentação adicional.**
            * Se você realizar correções (lógicas, de erros), adicionar documentação crucial que estava faltando, ou fizer melhorias significativas na estrutura, robustez ou clareza do código:
            * **Então, retorne o código Python MODIFICADO E APRIMORADO.**

        **O seu retorno deve ser exclusivamente o bloco de código Python (seja o original ou o modificado). Não inclua nenhuma introdução, explicação das mudanças, ou qualquer outro texto além do próprio código.**
        """
    )
    codigo_revisado = call_agent(buscador, codigo)
    return codigo_revisado

# --- Funções Auxiliares para Melhor Apresentação ---

def exibir_cabecalho_secao(titulo):
    """Exibe um cabeçalho formatado para destacar seções do processo."""
    print("\n" + "=" * 70)
    print(f"===== {titulo.upper()} =====")
    print("=" * 70 + "\n")

def solicitar_caminho_arquivo_video():
    """Solicita ao usuário o caminho do arquivo de vídeo e valida sua existência."""
    caminho_sugerido = "/content/drive/MyDrive/teste_ia.mp4"
    while True:
        path_video_input = input(f"Informe o caminho completo do arquivo de vídeo (ex: {caminho_sugerido}): ")
        if not path_video_input:
            print(f"Usando caminho padrão/sugerido: {caminho_sugerido}")
            path_video_input = caminho_sugerido 
        
        if os.path.exists(path_video_input):
            print(f"Arquivo de vídeo encontrado em: {path_video_input}")
            return path_video_input
        else:
            print(f"ERRO: Arquivo de vídeo não localizado em '{path_video_input}'. Por favor, verifique o caminho e tente novamente.")

def exibir_codigo_gerado(titulo_etapa, codigo_str):
    """Exibe o código gerado com um título para a etapa."""
    print(f"\n--- {titulo_etapa} ---")
    if hasattr(__builtins__, 'display'):
        display(Markdown(str(codigo_str)))
    else:
        print(str(codigo_str))
    print("-" * 70)

# --- Início da Execução Principal ---

exibir_cabecalho_secao("Agente de Análise de Vídeo e Geração de Código RPA")

resposta_necessidade_arq_extra = input("Além do vídeo principal, será necessário algum outro arquivo de entrada para o processo (ex: planilhas de dados)? (S/N): ")

try:
    path_video = solicitar_caminho_arquivo_video()

    exibir_cabecalho_secao("Geração Primária do Código a partir do Vídeo")
    codigo_bruto = verificar_video(path_video, Model_ID)
    exibir_codigo_gerado("Código Gerado Primariamente", codigo_bruto)

    exibir_cabecalho_secao("Validação de Necessidade de Arquivos/Métodos de Entrada Adicionais")
    resultado_processamento_entrada = arq_entrada_necessario(resposta_necessidade_arq_extra)
    exibir_codigo_gerado("Resultado do Processamento de Entradas Adicionais", resultado_processamento_entrada)

    exibir_cabecalho_secao("Revisão e Refinamento do Código")
    codigo_revisado_final = revisor_codigo_rpa(codigo_bruto, Model_ID)
    exibir_codigo_gerado("Código Final Após Revisão", codigo_revisado_final)
    
    print("\n\n>>> Processo de geração e revisão de código concluído! <<<")

except FileNotFoundError as e:
    print(f"\nERRO: Arquivo essencial não encontrado.")
    print(f"Detalhes: {e}")
except Exception as e:
    print(f"\nERRO INESPERADO DURANTE A EXECUÇÃO:")
    print(f"Tipo de Erro: {type(e).__name__}")
    print(f"Mensagem: {e}")
    print("O processo foi interrompido.")
